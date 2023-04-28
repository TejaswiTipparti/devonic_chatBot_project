from tkinter import *
import pandas as pd
import openpyxl
import re
import webbrowser
import spacy
import time
from tkinter import messagebox
from spacy.tokens import Token
from spellchecker import SpellChecker

nlp = spacy.load("en_core_web_sm")

BG_GRAY = "#ABB2B9"
BG_COLOR = "#17202A"
TEXT_COLOR = "#EAECEE"
FONT = "Helvetica 14"
FONT_BOLD = "Helvetica 13 bold"


brand_match = None
ram_range_match = None
hard_drive_match = None
price_start = None
price_end = None
brand = None
memory_size = None
hard_drive_size = None
hard_drive = None
price = None
price1 = None
price2 = None
ram = None
has_brand = False
has_price = False
has_ram = False
has_hard_drive = False
isAbove = False
isLess = False
deviceIdentificationSheetName = None
filtered_data = None
worksheet = None
msg1 = None


# workbook = openpyxl.load_workbook('static_data_collections.xlsx')
workbook = openpyxl.load_workbook('data.xlsx')

# Load the laptops data from Excel into a DataFrame
df = None

# create a SpellChecker instance
spell = SpellChecker()

# create a list of words to ignore
ignore_list = ["HP"]

# create a custom dictionary of words to ignore
custom_dict = set(ignore_list)

# add the custom dictionary to the SpellChecker instance
spell.word_frequency.load_words(custom_dict)

# Get all sheet names
sheet_names = workbook.sheetnames

# define tags for left and right messages
left_tag = "left"
right_tag = "right"


def correct_sentence(sentence):
    """
    This function takes a sentence as input, checks for misspelled words using the `spell` object (which is an instance
    of the `SpellChecker` class from the `pyspellchecker` library), corrects the misspelled words, and returns the
    corrected sentence.

    Parameters:
    - sentence: a string containing the sentence to be corrected

    Returns:
    - a string containing the corrected sentence
    """
    misspelled = spell.unknown(sentence.split())

    # create a list of corrected words
    corrected_words = []
    for word in sentence.split():
    # check if the word is misspelled
        if word in misspelled:
        # append the corrected word
            corrected_words.append(spell.correction(word))
        else:
        # append the original word
            corrected_words.append(word)

    # join the corrected words into a string
    corrected_text = " ".join(corrected_words)
    return corrected_text


class ChatApplication:
    """
    A class representing a chatbot that can interact with users through a graphical user interface (GUI).

    Methods:
    - __init__(): initializes the chatbot object by creating a Tkinter window, setting up the main window, 
      and scheduling the _check_idle() function to be called every 1000 milliseconds.
    - run(): starts the main loop of the Tkinter window, allowing the chatbot to interact with users.
    """
    def __init__(self):
        """
        Initializes the chatbot object by creating a Tkinter window, setting up the main window, 
        and scheduling the _check_idle() function to be called every 1000 milliseconds.
        """
        self.window = Tk()
        self._setup_main_window()
        self.window.after(1000, self._check_idle)
        self.last_activity = time.time()
        self._check_idle()
        
        
    def run(self):
        """
        Starts the main loop of the Tkinter window, allowing the chatbot to interact with users.
        """
        self.window.mainloop()

    # This function sets up the main window for the chatbot
    def _setup_main_window(self):
        # Set the window title, size, and background color
        self.window.title("DEVONIC")
        self.window.resizable(width=False, height=False)
        self.window.configure(width=470, height=530, bg=BG_COLOR)
        
        # Create a divider line

        line = Label(self.window, width=450, bg=BG_GRAY)
        line.place(relwidth=1, rely=0.07, relheight=1)
        
        # Create a text widget for displaying chat messages
        self.text_widget = Text(self.window, width=20, height=2, bg=BG_COLOR, fg=TEXT_COLOR,
                                font=FONT, padx=5, pady=5)
        self.text_widget.place(relheight=0.875, relwidth=0.975, rely=0)
        self.text_widget.configure(cursor="arrow", state=DISABLED)
        self.text_widget.tag_configure(left_tag, justify=LEFT)
        self.text_widget.tag_configure(right_tag, justify=RIGHT)

        # Add a welcome message to the text widget

        msg1 = "Hi, I am Devonic. I am here to get you best deals for Gadgets which you are looking for....\nPlease choose or type your requirement\n\n"
        self.text_widget.configure(state=NORMAL)
        self.text_widget.insert(END, msg1, left_tag)
        self.text_widget.configure(state=DISABLED)
        # Call a function to perform a common loop function with a list of sheet names

        self.perform_commom_Loop_function(sheet_names)
        # Add a blank line to the text widget

        msg2= "\n\n"
        self.text_widget.configure(state=NORMAL)
        self.text_widget.insert(END, msg2, left_tag)
        self.text_widget.configure(state=DISABLED)
        
        # Create a scrollbar for the text widget
        scrollbar = Scrollbar(self.window)
        scrollbar.pack(side=RIGHT, fill=Y)
        scrollbar.place(relx=0.975, relheight=1)
        scrollbar.configure(command=self.text_widget.yview)

        # Attach the scrollbar to the text widget
        self.text_widget.config(yscrollcommand=scrollbar.set)
        
        # Create a label at the bottom of the window
        bottom_label = Label(self.window, bg=BG_GRAY, height=80)
        bottom_label.place(relwidth=1, rely=0.900)
        
        # Create an entry box for the user to type messages
        self.msg_entry = Entry(bottom_label, bg="#2C3E50", fg=TEXT_COLOR, font=FONT)
        self.msg_entry.place(relwidth=0.74, relheight=0.03, rely=0.001, relx=0.011)
        self.msg_entry.focus()
        self.msg_entry.bind("<Return>", self._on_enter_pressed)
        
        # Create a send button for the user to send messages

        self.send_button = Button(bottom_label, text="Send", font=FONT_BOLD, width=20, bg=BG_GRAY,
                             command=lambda: self._on_enter_pressed(None))
        self.send_button.place(relx=0.77, rely=0.001, relheight=0.03, relwidth=0.22)
    
    # This method checks if the user is idle and prompts to restart the chatbot
    # if the user hasn't interacted with the chatbot for a certain amount of time
    def _check_idle(self):
        global brand, memory_size, hard_drive_size, price, price1, price2
        if self.text_widget.edit_modified() == 0 and self.msg_entry.get() == "":
            current_time = time.time()
            if current_time - self.last_activity > 300:
                self._disable_input()
                response = messagebox.askyesno("Restart Devonic ChatBot", "Do you want to restart the chatbot?")
                if response == 1:
                    self._enable_input()
                    self._insert_devonic_message("Chatbot restarted.\n Please select Laptop/ Mobile Option or else please provide input")
                    self.perform_commom_Loop_function(sheet_names)
                    brand = None
                    memory_size = None
                    hard_drive_size = None
                    price = None
                    price1 = None
                    price2 = None
                    deviceIdentificationSheetName = None
                    self.last_activity = time.time()
                    self.text_widget.edit_modified(False) # reset edit_modified flag
                    self.window.after(1000, self._check_idle)
                else:
                    self.window.destroy()
            else:
                self.window.after(1000, self._check_idle)
        else:
            self.last_activity = time.time()
            self.text_widget.edit_modified(False) # reset edit_modified flag
            self.window.after(1000, self._check_idle)

    
    def _on_enter_pressed(self, event):
        """
        This function is called when the user presses the "Enter" key while typing a message in the text entry widget. 
        It retrieves the text entered by the user, and inserts it as a message in the chat interface.

        Parameters:
        - self: the object that the function is called on (a ChatBot object in this case)
        - event: the event object that triggered the function (an "Enter" key press event in this case)
        """
        msg = self.msg_entry.get()
        self._insert_message(msg, "You")

    
    def _insert_message(self, msg, sender):
        """
        Inserts a message into the chat window, with the specified sender name and message text.

        Parameters:
        - msg: the text of the message to insert
        - sender: the name of the sender (i.e., the person or entity who sent the message)

         Returns: None
        """
        if not msg:
            return
        self.msg_entry.delete(0, END)
        msg1 = f"{sender}: {msg}\n\n"
        self.text_widget.configure(state=NORMAL)
        self.text_widget.insert(END, msg1, right_tag)
        self.text_widget.configure(state=DISABLED)
        self.get_response(msg)
        self.text_widget.see(END)
    
    def _insert_devonic_message(self,message):
        """
        This function is used to insert a message from the Devonic ChatBot into the text widget.

        Parameters:
        - self: the object that the function is called on (a ChatBot object in this case)
        - message: the message string to be inserted

        Output: None
        """
        msg2 = f"Devonic: {message}\n\n"
        self.text_widget.configure(state=NORMAL)
        self.text_widget.insert(END, msg2,left_tag)
        self.text_widget.configure(state=DISABLED)
        self.text_widget.see(END)
   
    def sheet_pressed(self, event, button_frame):
        """
            This function is called when a button is pressed in the GUI. It sets the global variable
            deviceIdentificationSheetName to the name of the sheet that was pressed, and then displays a
            message asking the user to choose a brand. It also creates a set of unique brands from the
            selected sheet, and then creates a row of buttons for every two unique brands, with each button
            corresponding to a brand. Finally, it inserts the button rows into the GUI.
    """
        global deviceIdentificationSheetName, worksheet
        deviceIdentificationSheetName = event.widget["text"]
        sheet_name = event.widget["text"]
        message1 = f"Choose a {deviceIdentificationSheetName} brand\n"
        self._insert_devonic_message(message1)

        worksheet = workbook[sheet_name]
        unique_brands = set()
        # Find the column index of the first occurrence of 'Brand' in the first row
        brand_col = None
        for cell in worksheet[1]:
            if cell.value == 'Brand':
                brand_col = cell.column
                break
            # Iterate over the rows in the worksheet, starting from the second row
        for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Extract the brand name from the row
            brand = row[brand_col - 1] # assuming 'Brand' is in column L (11th column)
                # Add the brand to the set of unique brands
            unique_brands.add(brand)

        button_rows = []
        unique_brands.remove(None)
        # convert the set of unique brands to a list
        unique_brands_list = list(unique_brands)

            # create a row of buttons for every two unique brands
        for i in range(0, len(unique_brands_list), 2):
                # create a new row frame
            row_frame = Frame(self.text_widget, bg=BG_COLOR)

                # add two buttons to the current row frame
            for brand in unique_brands_list[i:i+2]:
                button = Button(row_frame, text=brand, bg=BG_COLOR)
                button.pack(side=LEFT, padx=5, pady=5)
                button.bind("<Button-1>", self.brand_button_click)

                # add the row frame to the list of button rows
            button_rows.append(row_frame)

        # insert all the button rows into the text widget
        for row_frame in button_rows:
            self.text_widget.window_create(END, window=row_frame)
            self.text_widget.configure(state=NORMAL)
            self.text_widget.insert(END, "\n")
            self.text_widget.configure(state=DISABLED)

        self.text_widget.see(END)

    def brand_button_click(self, event):
        # declare global variables that will be used to store the device sheet name, brand, and data frame
        global deviceIdentificationSheetName, brand, df

       # get the text of the button that was clicked

        tag_name = event.widget["text"]
        brand = tag_name
        # read in the data from the Excel file for the specified device sheet name

        df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)

        # select the columns that contain memory information
        memory_cols = [col for col in df.columns if 'Memory' in col]
        # filter the data frame to only include rows with the selected brand

        filtered_df = df[df["Brand"] == brand]
        # select the related memory columns for the filtered rows
        selected_cols = ["Brand"] + memory_cols
        result_df = filtered_df[selected_cols]
        # get a list of unique RAM sizes from the filtered data frame
        ram_options = result_df["Memory"].unique().tolist()
        # create a message to display to the user
        message1 = f"The available RAM size for {brand} {deviceIdentificationSheetName} are"
        # insert the message into the chat window
        self._insert_devonic_message(message1)
        # add buttons to the current row frame
        button_rows1 = []
        for ram in ram_options:
            row_frame1 = Frame(self.text_widget, bg=BG_COLOR)
            button = Button(row_frame1, text=ram, bg=BG_COLOR)
            button.pack(side=LEFT, padx=5, pady=5)
            button.bind("<Button-1>", self.memory_button_click)
            button_rows1.append(row_frame1)
        # insert all the button rows into the text widget
        for row_frame in button_rows1:
            self.text_widget.window_create(END, window=row_frame)
            self.text_widget.configure(state=NORMAL)
            self.text_widget.insert(END, "\n")
            self.text_widget.configure(state=DISABLED)

        self.text_widget.see(END)
       

    def memory_button_click(self, event):
        """
        This function is called when the user clicks on a memory button. It reads data from an Excel file, filters the data based on the selected brand and memory size, and displays available options for either hard drive or price range, depending on the device type.

        Parameters:
        - self: the instance of the class that this method belongs to
        - event: the event object that triggered this function call

        Returns:
        - None
        """
        global deviceIdentificationSheetName, brand, df, memory_size
        tag_name = event.widget["text"]
        memory_size = tag_name
        df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)
       
        memory_cols = [col for col in df.columns if 'Memory' in col]
        filtered_df = df[(df['Brand'] == brand) & (df[memory_cols[0]].str.contains(memory_size, case=False))]
        if deviceIdentificationSheetName.lower() == 'laptop' :
            hard_drive_cols = [col for col in df.columns if 'Hard Drive' in col and col not in memory_cols]

            # select the related memory columns for the filtered rows
            selected_cols = ["Brand"] + hard_drive_cols
            result_df = filtered_df[selected_cols]

            hard_drive_options = result_df["Hard Drive"].unique().tolist()
            message1 = f"The available storage for an {brand} {deviceIdentificationSheetName} of {memory_size} RAM is"
            self._insert_devonic_message(message1)
            # add buttons to the current row frame
            button_rows1 = []
            for ram in hard_drive_options:
                row_frame1 = Frame(self.text_widget, bg=BG_COLOR)
                button = Button(row_frame1, text=ram, bg=BG_COLOR)
                button.pack(side=LEFT, padx=5, pady=5)
                button.bind("<Button-1>", self.hard_drive_button_click)
                button_rows1.append(row_frame1)
            # insert all the button rows into the text widget
            for row_frame in button_rows1:
                self.text_widget.window_create(END, window=row_frame)
                self.text_widget.configure(state=NORMAL)
                self.text_widget.insert(END, "\n")
                self.text_widget.configure(state=DISABLED)

            self.text_widget.see(END)
        elif deviceIdentificationSheetName.lower() == 'mobile':
            # Get the Price column
            price_options = filtered_df['Price'].unique().tolist()
            message1 = f"Please select The available options price range between ${filtered_df['Price'].min()} and ${filtered_df['Price'].max()}."
            self._insert_devonic_message(message1)
            filtered_data = df[(df['Brand'].str.lower() == brand.lower()) & (df['Memory'].str.contains(memory_size))]

            # Get the best price from Amazon and Best Buy
            amazon_data = filtered_data[filtered_data['Website'] == 'amazon']
            amazon_best_price = amazon_data['Price'].min()
            bestbuy_data = filtered_data[filtered_data['Website'] == 'bestbuy']
            bestbuy_best_price = bestbuy_data['Price'].min()

            # Determine which website has the better price and display a message to the user
            if amazon_best_price < bestbuy_best_price:
                message1 = f'Amazon has the better price at ${amazon_best_price:.2f}'
                self._insert_devonic_message(message1) 
            elif bestbuy_best_price < amazon_best_price:
                message1 = f'Bestbuy has the better price at ${bestbuy_best_price:.2f}'
                self._insert_devonic_message(message1) 
            else:
                message1 = f'Amazon and BestBuy have the same price with ${bestbuy_best_price:.2f}'
                self._insert_devonic_message(message1) 
            # add buttons to the current row frame
            button_rows1 = []
            for ram in price_options:
                row_frame1 = Frame(self.text_widget, bg=BG_COLOR)
                button = Button(row_frame1, text=ram, bg=BG_COLOR)
                button.pack(side=LEFT, padx=5, pady=5)
                button.bind("<Button-1>", self.price_button_click)
                button_rows1.append(row_frame1)
            # insert all the button rows into the text widget
            for row_frame in button_rows1:
                self.text_widget.window_create(END, window=row_frame)
                self.text_widget.configure(state=NORMAL)
                self.text_widget.insert(END, "\n")
                self.text_widget.configure(state=DISABLED)
            self.text_widget.see(END)

    # This function is called when the user clicks on a button for selecting a hard drive size
    def hard_drive_button_click(self, event):
        # Declare global variables that are used in this function
        global deviceIdentificationSheetName, brand, df, memory_size, hard_drive_size
        # Get the text of the button that was clicked and set it as the hard drive size

        tag_name = event.widget["text"]
        hard_drive_size = tag_name
        # Read data from an Excel file into a Pandas DataFrame

        df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)
        # Filter the DataFrame to only include rows that match the selected brand, memory size, and hard drive size

        memory_cols = [col for col in df.columns if 'Memory' in col]
        hard_drive_cols = [col for col in df.columns if 'Hard Drive' in col]
        price_cols = [col for col in df.columns if 'Price' in col]
        filtered_df = df[(df['Brand'] == brand) & (df[memory_cols[0]].str.contains(memory_size, case=False)) &(df[hard_drive_cols[0]].str.contains(hard_drive_size, case=False))]
        # Select specific columns from the filtered DataFrame to display to the user

        selected_cols = ["Brand"] + memory_cols + hard_drive_cols + price_cols
        result_df = filtered_df[selected_cols]
        
        # Get a list of available price options and display a message to the user
        price_options = result_df["Price"].unique().tolist()
        message1 = f"Please select The available options price range between ${filtered_df['Price'].min()} and ${filtered_df['Price'].max()}."
        self._insert_devonic_message(message1) 
        # Filter the data again to only include rows that match the selected brand, memory size, and hard drive size
        filtered_data = df[(df['Brand'].str.lower() == brand.lower()) & (df['Memory'].str.contains(memory_size)) & (df['Hard Drive'].str.contains(hard_drive_size))]

        # Get the best price from Amazon and Best Buy
        amazon_data = filtered_data[filtered_data['Website'] == 'amazon']
        amazon_best_price = amazon_data['Price'].min()
        costco_data = filtered_data[filtered_data['Website'] == 'costco']
        costco_best_price = costco_data['Price'].min()
        # Determine which website has the better price and display a message to the user
        if amazon_best_price < costco_best_price:
            message1 = f'Amazon has the better price at ${amazon_best_price:.2f}'
            self._insert_devonic_message(message1) 
            print(f'Amazon has the better price at ${amazon_best_price:.2f}')
        elif costco_best_price < amazon_best_price:
            message1 = f'Costco has the better price at ${costco_best_price:.2f}'
            self._insert_devonic_message(message1) 
            print(f'Costco has the better price at ${costco_best_price:.2f}')
        else:
            message1 = f'Both Amazon and Costco have the same price with ${costco_best_price:.2f}'
            self._insert_devonic_message(message1) 
            print('Both Amazon and Costco have the same price')  
        # Add buttons for each available price option to the text widget
        button_rows1 = []
        for ram in price_options:
            row_frame1 = Frame(self.text_widget, bg=BG_COLOR)
            button = Button(row_frame1, text=ram, bg=BG_COLOR)
            button.pack(side=LEFT, padx=5, pady=5)
            button.bind("<Button-1>", self.price_button_click)
            button_rows1.append(row_frame1)
        for row_frame in button_rows1:
            self.text_widget.window_create(END, window=row_frame)
            self.text_widget.configure(state=NORMAL)
            self.text_widget.insert(END, "\n")
            self.text_widget.configure(state=DISABLED)
        self.text_widget.see(END)
             
    
    def price_button_click(self, event):
        """
        This function is called when a button is clicked. It filters data from an Excel sheet based on certain criteria 
        (brand, memory size, hard drive size, and price), and displays the filtered data in a text widget.

        Parameters:
        - self: the object that the function is called on (a ChatBot object in this case)
        - event: the event object that triggered the function (a button click event in this case)
        """
        global deviceIdentificationSheetName, brand, df, memory_size, hard_drive_size, price
        tag_name = event.widget["text"]
        price = event.widget["text"]
        
        if deviceIdentificationSheetName.lower() == 'laptop':
            df = pd.read_excel('data.xlsx', sheet_name=deviceIdentificationSheetName)
            memory_cols = [col for col in df.columns if 'Memory' in col]
            price_cols = [col for col in df.columns if 'Price' in col]
            hard_drive_cols = [col for col in df.columns if 'Hard Drive' in col]
        
            filtered_df = df[(df['Brand'].str.lower() == brand.lower()) & 
                 (df[memory_cols[0]].str.contains(memory_size, case=False)) &
                 (df[hard_drive_cols[0]].str.contains(hard_drive_size, case=False)) &
                 (df[price_cols[0]] == price)]
            # retrieve the entire rows for the filtered rows
            result_df = filtered_df.reset_index(drop=True)

            row_frame = Frame(self.text_widget, bg=BG_COLOR)
            self.text_widget.tag_configure('hyperlink', foreground='blue', underline=True)
            for index, row in filtered_df.iterrows():
                url = row['Url']
                hyperlink_text = f"Click here to more details {row['Website']}"
                self.text_widget.window_create(END, window=row_frame)
                self.text_widget.configure(state=NORMAL)
                self.text_widget.insert(END, f"\n{row['Product Name']}\n")
                self.text_widget.insert(END, hyperlink_text, 'hyperlink')
                self.text_widget.insert(END, f"\nPrice: {row['Price']}\n")
                self.text_widget.tag_bind('hyperlink', '<Button-1>', lambda event, url=url: webbrowser.open(url, new=2, autoraise=True))
                self.text_widget.insert(END, '\n')
                self.text_widget.configure(state=DISABLED)
            self.text_widget.see(END)
        elif deviceIdentificationSheetName.lower() == 'mobile':
            df = pd.read_excel('data.xlsx', sheet_name=deviceIdentificationSheetName)
            memory_cols = [col for col in df.columns if 'Memory' in col]
            price_cols = [col for col in df.columns if 'Price' in col]
            filtered_df = df[(df['Brand'].str.lower() == brand) & 
                 (df[memory_cols[0]].str.contains(memory_size, case=False)) &
                 (df[price_cols[0]] == price)]

            row_frame = Frame(self.text_widget, bg=BG_COLOR)
            self.text_widget.tag_configure('hyperlink', foreground='blue', underline=True)
            for index, row in filtered_df.iterrows():
                # #print each column value of the current row
                url = row['Url']
                hyperlink_text = f"Click here to more details {row['Website']}"
                self.text_widget.window_create(END, window=row_frame)
                self.text_widget.configure(state=NORMAL)
                self.text_widget.insert(END, f"\n{row['Product Name']}\n")
                self.text_widget.insert(END, hyperlink_text, 'hyperlink')
                self.text_widget.insert(END, f"\nPrice: {row['Price']}\n")
                self.text_widget.tag_bind('hyperlink', '<Button-1>', lambda event, url=url: webbrowser.open_new(url))
                self.text_widget.insert(END, '\n')
                self.text_widget.configure(state=DISABLED)
            self.text_widget.see(END)
        self._insert_devonic_message("Thank you for Devonic ChatBot")

    def perform_commom_Loop_function(self,list_options):
        """
        This function creates a frame widget to hold buttons for each sheet name in the list_options. 
        It then creates buttons inside the frame widget and binds the button click event to a sheet_pressed method. 
        Finally, it inserts the button frame into the text widget and moves the cursor to the end of the widget.

        Args:
            self: The instance of the class containing this method.
            list_options: A list of sheet names.

        Returns:
            None
        """
        # Create a Frame widget to hold the buttons
        button_frame = Frame(self.text_widget,bg=BG_COLOR)
        button_frame.pack(side=TOP, padx=5, pady=5)
        button_frame.config(bg=BG_COLOR)
        # Create buttons inside the button_frame
        for sheet_name in list_options:
            button1 = Button(button_frame, text=sheet_name,bg=BG_COLOR)
            # Pack the buttons in the button_frame
            button1.pack(side=LEFT, padx=5, pady=5)
            # Bind the button click event to a method
            button1.bind("<Button-1>", lambda event:self.sheet_pressed(event, button_frame))
        # Insert the button_frame into the Text widget
        self.text_widget.window_create("end", window=button_frame)
        self.text_widget.insert(END, '\n')
        self.text_widget.see(END) 

    def _disable_input(self):
        """
        This method disables the message entry and send button of the chat interface, so that the user cannot send any 
        messages while the bot is processing a request or sending a response.
    
        Parameters:
         - self: the object that the method is called on (a ChatBot object in this case)
        """
        self.msg_entry.configure(state=DISABLED)
        self.send_button.configure(state=DISABLED)
        
    def _enable_input(self):
        """
        This method enables the message entry and send button of the chat interface, so that the user can send messages 
        again after they were disabled.
    
        Parameters:
            - self: the object that the method is called on (a ChatBot object in this case)
         """
        self.msg_entry.configure(state=NORMAL)
        self.send_button.configure(state=NORMAL)
    
    def get_response(self, message):
        """
        Processes a message and returns a response.

        Args:
            message (str): The message to be processed.

        Returns:
            str: The response to the message.

        Raises:
            None

        Description:
            This function takes a message as input and uses natural language processing to extract
        information from it, such as the type of device (laptop or mobile) and any relevant specifications
        (such as brand, RAM, or hard drive size). It then reads in data from an Excel file and filters
        the data based on the extracted information. Finally, it generates a response based on the filtered
        data.
        """
        global brand_match,df,filtered_data, deviceIdentificationSheetName, price_end, price_start,ram_range_match, hard_drive_match,price, price1, price2, brand,ram, hard_drive, has_brand, has_hard_drive,has_price, has_ram, isAbove, isLess,memory_size,hard_drive_size
        text = correct_sentence(message)
        doc = nlp(text)
        tokens = []
        for token in doc:
            if token.text.lower() in ['laptop', 'notebook', 'macbook', 'chromebook']:
                deviceIdentificationSheetName ='Laptop'
            elif token.text.lower() in ['mobile', 'smartphone', 'iphone', 'android']:
                deviceIdentificationSheetName = 'Mobile'
        df = pd.read_excel("data.xlsx", deviceIdentificationSheetName)
        for token in doc:
            tokens.append(token.text)
        if deviceIdentificationSheetName is None:
            # If device identification sheet name is None, request the user to try retrieving products as a laptop or mobile.
            self._insert_devonic_message("Please try as laptop or mobile for reterive Products")
        else:
            # Get unique values from the 'Brand' column of the input DataFrame
            unique_values = df['Brand'].dropna()[df['Brand'] != 'None'].unique()
            # Create a regular expression pattern for brand matching
            pattern = r"\b(" + '|'.join(map(re.escape, unique_values)) + r")\b"
            # Compile regular expression with ignore case flag
            brand_regex = re.compile(pattern, flags=re.IGNORECASE)
            # Create a regular expression pattern for RAM matching
            ram_regex = re.compile(r"\b(\d+)\s*(?:GB|gb)\s*(?:RAM|ram)\b")
            # pricePattern = r"(?:above|over|more than|between|less than)?[^\d]*(\d+(?:,\d{3})*(?:\.\d+)?)"
            # pricePattern = r"(?:above|over|more than|between|less than)?[^\d]*(\d+(?:,\d{3})*(?:\.\d+)?)(?!\s*[TtGg]\b|\s*\d\s*[Gg][Bb]\b)"
            # Create a regular expression pattern for price matching
            pricePattern = r"(?:above|over|more than|between|less than)?[^\d$]*(?<![\d$])(\$\d+(?:,\d{3})*(?:\.\d+)?|\d+(?:,\d{3})*(?:\.\d+)?\$?)(?!\s*[TtGg]\b| GB| TB|\d+\s*(?:GB|TB))"

            # Find all price matches in the input message using the price pattern
            matches = re.findall(pricePattern, message, re.IGNORECASE)
            # Find a brand match in the input message using the brand regex pattern
            brand_match = brand_regex.search(message)
            # Find a RAM range match in the input message using the RAM regex pattern
            ram_range_match = ram_regex.search(message, re.IGNORECASE)

            if deviceIdentificationSheetName.lower() == 'laptop':
                # Find a hard drive match in the input message using the hard drive regex pattern
                hard_drive_regex = re.compile(r'(\d+)\s*(?:TB|GB)\s*(?:\+)?\s*(\d+)?\s*(?:TB|GB)?\s*(?:hard drive|HDD|solid state drive|SSD)', flags=re.IGNORECASE)
                hard_drive_match = hard_drive_regex.search(message)
        
            if brand_match:
                # Loop through each token in the input message and check if it matches with the brand regex pattern
                for token in doc:
                    if re.search(brand_regex, token.text):
                        # If a match is found, convert the matched brand name to lowercase and set has_brand to True

                        brand = token.text.lower()
                        has_brand = True

            if ram_range_match:
            # #print the list of tokens
                tokens1 = []
            # Loop through the doc object and append each token to the list
                for token in doc:
                    tokens1.append(token.text)
                text = ""
                ram_size = ""
            # Loop through tokens
                for i, token in enumerate(tokens):
                # If the token represents a RAM size (e.g. "8GB"), add it to the text string
                    if token.isdigit() and tokens[i+1].lower() == "gb" and text == "":
                        text = token + "" + tokens[i+1]
                # If the token represents "RAM", add it to the text string
                    elif token.lower() == "ram" and text != "":
                        text += "" + token
                # If we have a complete RAM specification, search for it using the regular expression pattern
                    if ram_regex.search(text):
                        match = ram_regex.search(text)
                        ram_size = match.group()
                        ram_size = ram_size[:-3]
                    # Reset the text string
                        text = ""
                if ram_size != "":
                    has_ram = True
                    memory_size = ram_size

            if deviceIdentificationSheetName.lower() == 'laptop' and hard_drive_match:
                # Initialize variable
                text = ""
                hard_drive_size = ""
                # Loop through tokens in the document
                for token in doc:
                    # If the token represents a hard drive size (e.g. "1TB"), add it to the text string
                    if token.pos_ == "NUM" and token.tag_ == "CD":
                        text = token.text
                    elif token.text.lower() in ["tb", "gb"] and text != "":
                        text += "" + token.text
                    # If the token represents "hard drive", "HDD", "solid state drive", or "SSD", add it to the text string
                    elif token.text.lower() in ["hard", "drive", "hdd", "solid", "state", "ssd"] and text != "":
                        text += " " + token.text
                    # If we have a complete hard drive specification, search for it using the regular expression pattern
                    if hard_drive_regex.search(text):
                        match = hard_drive_regex.search(text)
                        hard_drive_size = match.group()
                        # Reset the text string
                        hard_drive_size = hard_drive_size[:-10]
                        text = ""

                if hard_drive_size != "":
                    has_hard_drive = True

            if matches:
                # If there are one or more matches for a price in the message:
                if len(matches) >= 2:
                    # If there are two or more matches, assume that the lower price is the first match and the higher price is the second match:
                    price1 = float(matches[0].replace(",", ""))
                    price2 = float(matches[1].replace(",", ""))
                    if price1 < price2:
                        # If the first price is lower than the second price, set the "has_price" flag to True:
                        has_price = True
                    else:
                        # If the second price is lower than or equal to the first price, set both prices to None:
                        price1 = None
                        price2 = None
                elif len(matches) == 1:
                    # If there is only one match, assume that it represents the desired price:
                    price = float(matches[0].replace(",", ""))
                    if "less" in message.lower():
                        # If the message contains the word "less", set the "has_price" and "isLess" flags to True:
                        has_price = True
                        isLess = True
                        isAbove = False
                    elif "above" in message.lower() or "over" in message.lower() or "more than" in message.lower():
                        # If the message contains the words "above", "over", or "more than", set the "has_price" and "isAbove" flags to True:
                        has_price = True
                        isAbove = True
                        isLess = False
                    else:
                        # If the message does not contain any of the above words, set the price to None and both flags to False:
                        price = None
                        isAbove = False
                        isLess = False
             
        
            if has_brand is False:
                # Display message to prompt user to select a brand from the available options
                message1 = f"{deviceIdentificationSheetName} Brand name is unidentified from the input.\mHere are the list of Brand Name available near us {', '.join(unique_values)}. \nPlease check a Brand name from the list given here. "
                self._insert_devonic_message(message1)
            elif has_ram is False:
                # If brand is provided but RAM size is missing, display available RAM options for the selected brand
                df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)
                memory_cols = [col for col in df.columns if 'Memory' in col]
                filtered_df = df[df["Brand"].str.lower() == brand]

                # select the related memory columns for the filtered rows
                selected_cols = ["Brand"] + memory_cols
                result_df = filtered_df[selected_cols]

                ram_options = result_df["Memory"].unique().tolist()
                message1 = f"{brand.upper()} {deviceIdentificationSheetName} RAM Details are Missing. Please check here are the available RAM options are: {', '.join(ram_options)} and try to keep RAM at the end of the size.\nFor eg: 12GB RAM"
                self._insert_devonic_message(message1)
            # Check if the device identification sheet name is "laptop" and if the device has a hard drive
            elif deviceIdentificationSheetName.lower() == 'laptop'and has_hard_drive is False:
                # Read the data from the excel file
                df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)
                # Get the columns containing memory information
                memory_cols = [col for col in df.columns if 'Memory' in col]
                # Filter the dataframe based on the given brand and memory size
                filtered_df = df[(df['Brand'].str.lower() == brand) & (df[memory_cols[0]].str.contains(memory_size, case=False))]
                # If the filtered dataframe is empty, retrieve all the devices from the given brand and provide a list of available RAM options
                if filtered_df.empty:
                    filtered_df = df[df["Brand"].str.lower() == brand]
                    ram_selected_cols = ["Brand"] + memory_cols
                    ram_result_df = filtered_df[ram_selected_cols]
                    ram_options = ram_result_df["Memory"].unique().tolist() 
                    has_ram = False
                    message1 = f"Sorry!! We didn't find any results for {brand.upper()} {deviceIdentificationSheetName} with {memory_size} RAM. Please check here are the available RAM options are: {', '.join(ram_options)}."
                    self._insert_devonic_message(message1)
                # Otherwise, retrieve all the available hard drive options for the filtered devices and provide a message to the user to check the available hard drive options and select one
                else:
                    hard_drive_cols = [col for col in df.columns if 'Hard Drive' in col and col not in memory_cols]
            
                    selected_cols = ["Brand"] + hard_drive_cols
                    result_df = filtered_df[selected_cols]
                    hard_drive_options = result_df["Hard Drive"].unique().tolist()
                    message1 = f"{brand.upper()} {deviceIdentificationSheetName} with {memory_size} RAM of Hard Drive Details are missing. Please check here are the available Hard Drive options are: {', '.join(hard_drive_options)} and try to keep Hard drive at the end of the size.\n For eg: 1TB Hard drive" 
                    self._insert_devonic_message(message1)
            elif has_price is False:
                # This code provides a search functionality for laptop and mobile devices based on brand, memory size, and hard drive size (for laptops only), and displays the available price ranges for the selected device.
                if deviceIdentificationSheetName.lower() == 'laptop':
                    # Read the data from the Excel sheet
                    df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)
                    # Get the columns that contain the memory and hard drive information for laptops
                    memory_cols = [col for col in df.columns if 'Memory' in col]
                    hard_drive_cols = [col for col in df.columns if 'Hard Drive' in col and col not in memory_cols]
                    price_cols = [col for col in df.columns if 'Price' in col]
                    # Filter the data based on the selected brand, memory size, and hard drive size (if applicable)
                    filtered_df = df[(df['Brand'].str.lower() == brand) & (df[memory_cols[0]].str.contains(memory_size, case=False)) & (df[hard_drive_cols[0]].str.contains(hard_drive_size, case=False))]
                    # If no results were found, display a message with the available hard drive options
                    if filtered_df.empty:
                        has_hard_drive = False
                        filtered_df = df[(df['Brand'].str.lower() == brand) & (df[memory_cols[0]].str.contains(memory_size, case=False))]
                        selected_cols = ["Brand"] + hard_drive_cols
                        hard_result_df = filtered_df[selected_cols]
                        hard_drive_options = hard_result_df["Hard Drive"].unique().tolist()
                        message1 = f"Sorry!! We didn't find any results for {brand.upper()} {deviceIdentificationSheetName} of {memory_size} RAM with {hard_drive_size} storage. Please check here are the available storage options are: {', '.join(hard_drive_options)}."
                        self._insert_devonic_message(message1)
                    # If results were found, display the available price ranges and create buttons for selecting a range
                    else:
                        selected_cols = ["Brand"] + memory_cols + hard_drive_cols + price_cols
                        result_df = filtered_df[selected_cols]
                        price_options = result_df["Price"].unique().tolist()
                        message1 = f"{brand.upper()} {deviceIdentificationSheetName} of {memory_size} RAM with {hard_drive_size} storage wher price Range is mssing. Please give input value with less or above or between to.\n Here are The available options price range between ${filtered_df['Price'].min()} and ${filtered_df['Price'].max()}."
                        self._insert_devonic_message(message1)
                        button_rows1 = []
                        # Create buttons inside the button_frame
                        for ram in price_options:
                            row_frame1 = Frame(self.text_widget, bg=BG_COLOR)
                            button = Button(row_frame1, text=ram, bg=BG_COLOR)
                            button.pack(side=LEFT, padx=5, pady=5)
                            button.bind("<Button-1>", self.price_button_click)
                            button_rows1.append(row_frame1)
                        # insert all the button rows into the text widget
                        for row_frame in button_rows1:
                            self.text_widget.window_create(END, window=row_frame)
                            self.text_widget.configure(state=NORMAL)
                            self.text_widget.insert(END, "\n")
                            self.text_widget.configure(state=DISABLED)
                        self.text_widget.see(END)
                elif deviceIdentificationSheetName.lower() == 'mobile':
                    # Read the data from the Excel sheet
                    df = pd.read_excel('data.xlsx', sheet_name = deviceIdentificationSheetName)
                    # Get the columns that contain the memory for mobiles
                    memory_cols = [col for col in df.columns if 'Memory' in col]
                    # Filter the dataframe based on the given brand and memory size
                    filtered_df = df[(df['Brand'].str.lower() == brand) & (df[memory_cols[0]].str.contains(memory_size, case=False))]
                    price_options = filtered_df['Price'].unique().tolist()

                    message1 = f"Price Range is mssing. Please try selecting are the available options prices ${filtered_df['Price'].min()} and ${filtered_df['Price'].max()}."
                    self._insert_devonic_message(message1) 
                    # Create buttons inside the button_frame
                    button_rows1 = []
                    for ram in price_options:
                        row_frame1 = Frame(self.text_widget, bg=BG_COLOR)
                        button = Button(row_frame1, text=ram, bg=BG_COLOR)
                        button.pack(side=LEFT, padx=5, pady=5)
                        button.bind("<Button-1>", self.price_button_click)
                        button_rows1.append(row_frame1)
                    # insert all the button rows into the text widget
                    for row_frame in button_rows1:
                        self.text_widget.window_create(END, window=row_frame)
                        self.text_widget.configure(state=NORMAL)
                        self.text_widget.insert(END, "\n")
                        self.text_widget.configure(state=DISABLED)
                    self.text_widget.see(END)


if __name__ == "__main__":
    app = ChatApplication()
    app.run()
    
    